from enums import Emotions
import openpyxl
from openpyxl.styles import Font


class ExcelExporter:
    def __init__(self, file_name):
        self.filename = file_name
        workbook = openpyxl.Workbook()
        worksheet = workbook.get_active_sheet()

        # Write headers
        worksheet['A1'] = 'File'
        worksheet['B1'] = 'Number of Faces'
        worksheet['C1'] = 'Camera Instability'
        worksheet['D1'] = 'Detection Confidence'
        worksheet['E1'] = 'Head Pose - Roll'
        worksheet['F1'] = 'Head Pose - Pan'
        worksheet['G1'] = 'Head Pose - Tilt'

        # # Google API
        # worksheet['H1'] = 'Emotions - Joy'
        # worksheet['I1'] = 'Emotions - Sorrow'
        # worksheet['J1'] = 'Emotions - Anger'
        # worksheet['K1'] = 'Emotions - Surprised'

        # Azure API
        worksheet['H1'] = 'Emotions - Surprise'
        worksheet['I1'] = 'Emotions - Happiness'
        worksheet['J1'] = 'Emotions - Fear'
        worksheet['K1'] = 'Emotions - Disgust'
        worksheet['L1'] = 'Emotions - Neutral'
        worksheet['M1'] = 'Emotions - Anger'
        worksheet['N1'] = 'Emotions - Sadness'
        worksheet['O1'] = 'Emotions - Contempt'
        worksheet['P1'] = 'Smile'
        worksheet['Q1'] = 'Score'

        # Set headers to bold
        bold_font = Font(bold=True)
        for cell in worksheet["1:1"]:
            cell.font = bold_font

        # Save file
        workbook.save(self.filename)

        # Set append_row to first line after headers
        self.append_row = 2

    def append_results(self, results, score, file):
        # Load the file
        results_file = openpyxl.load_workbook(self.filename)

        # Get active sheet
        results_sheet = results_file.get_active_sheet()

        # Write values in row
        results_sheet.cell(row=self.append_row, column=1).value = file
        results_sheet.cell(row=self.append_row, column=2).value = results.face_count
        results_sheet.cell(row=self.append_row, column=3).value = results.camera_instability
        results_sheet.cell(row=self.append_row, column=4).value = results.detection_confidence
        results_sheet.cell(row=self.append_row, column=5).value = results.head_pose['roll_angle']
        results_sheet.cell(row=self.append_row, column=6).value = results.head_pose['pan_angle']
        results_sheet.cell(row=self.append_row, column=7).value = results.head_pose['tilt_angle']

        # Google API
        # results_sheet.cell(row=self.append_row, column=8).value = results.emotions[Emotions.happiness]
        # results_sheet.cell(row=self.append_row, column=9).value = results.emotions[Emotions.sorrow]
        # results_sheet.cell(row=self.append_row, column=10).value = results.emotions[Emotions.anger]
        # results_sheet.cell(row=self.append_row, column=11).value = results.emotions[Emotions.surprise]

        # Azure API
        results_sheet.cell(row=self.append_row, column=8).value = results.emotions['surprise']
        results_sheet.cell(row=self.append_row, column=9).value = results.emotions['happiness']
        results_sheet.cell(row=self.append_row, column=10).value = results.emotions['fear']
        results_sheet.cell(row=self.append_row, column=11).value = results.emotions['disgust']
        results_sheet.cell(row=self.append_row, column=12).value = results.emotions['neutral']
        results_sheet.cell(row=self.append_row, column=13).value = results.emotions['anger']
        results_sheet.cell(row=self.append_row, column=14).value = results.emotions['sadness']
        results_sheet.cell(row=self.append_row, column=15).value = results.emotions['contempt']
        results_sheet.cell(row=self.append_row, column=16).value = results.smile
        results_sheet.cell(row=self.append_row, column=17).value = score

        # Increment append_row
        self.append_row += 1

        # Save file
        results_file.save(self.filename)
